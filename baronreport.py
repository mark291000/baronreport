import streamlit as st
import pandas as pd
import openpyxl
import base64
from datetime import datetime
import plotly.graph_objects as go
import plotly.express as px

# =========================================================
# TASK DASHBOARD - STREAMLIT VERSION
# =========================================================

# === Cấu hình trang ===
st.set_page_config(
    page_title="Task Dashboard",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded"
)

# === CSS tùy chỉnh ===
st.markdown("""
<style>
    .main {
        padding: 0rem 1rem;
    }
    .stDataFrame {
        width: 100%;
    }
    div[data-testid="stExpander"] div[role="button"] p {
        font-size: 1.1rem;
        font-weight: bold;
    }
    .status-completed {
        background-color: #d4edda;
        padding: 5px 10px;
        border-radius: 5px;
        font-weight: bold;
        color: #155724;
    }
    .status-working {
        background-color: #fff3cd;
        padding: 5px 10px;
        border-radius: 5px;
        font-weight: bold;
        color: #856404;
    }
    .status-delay {
        background-color: #f8d7da;
        padding: 5px 10px;
        border-radius: 5px;
        font-weight: bold;
        color: #721c24;
    }
    .status-newtask {
        background-color: #cce5ff;
        padding: 5px 10px;
        border-radius: 5px;
        font-weight: bold;
        color: #004085;
    }
</style>
""", unsafe_allow_html=True)

# === Hàm xử lý dữ liệu ===
@st.cache_data
def load_and_process_data(uploaded_file):
    """Load và xử lý dữ liệu từ Excel file"""
    today = pd.Timestamp.now().normalize()
    
    # Đọc Excel với openpyxl để lấy hình ảnh
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    ws = wb.active
    
    # Lưu hình ảnh (cột PICTURE)
    images = {}
    for image in ws._images:
        cell = image.anchor._from
        cell_coord = f"{openpyxl.utils.get_column_letter(cell.col + 1)}{cell.row + 1}"
        img_bytes = image.ref.getvalue() if hasattr(image.ref, 'getvalue') else image.ref
        images[cell_coord] = base64.b64encode(img_bytes).decode('utf-8')
    
    # Đọc Excel với pandas (header dòng 3)
    df = pd.read_excel(uploaded_file, header=2, engine='openpyxl')
    df.columns = df.columns.str.strip()
    
    # Xử lý Status
    def get_status(row):
        confirm = str(row.get("CONFIRM FROM BARON", "")).strip().lower()
        start_date = row.get("START DATE")
        due = row.get("DUE DATE")
        
        if "go" in confirm:
            return "Completed"
        
        if pd.notna(start_date):
            try:
                if pd.to_datetime(start_date).date() > today.date():
                    return "New Task"
            except:
                pass
        
        if pd.notna(due):
            try:
                if pd.to_datetime(due).date() < today.date():
                    return "Delay"
                else:
                    return "Working"
            except:
                return "Working"
        
        return "Working"
    
    df["STATUS"] = df.apply(get_status, axis=1)
    
    # Định dạng ngày
    df["START DATE"] = pd.to_datetime(df["START DATE"], errors="coerce")
    df["DUE DATE"] = pd.to_datetime(df["DUE DATE"], errors="coerce")
    
    # Thêm cột PICTURE_BASE64 từ images dict
    df["PICTURE_BASE64"] = ""
    for idx, row in df.iterrows():
        row_num = idx + 4  # Bắt đầu từ dòng 4 trong Excel
        # Tìm cột PICTURE
        header_row = [cell.value for cell in ws[3]]
        try:
            picture_col_idx = header_row.index("PICTURE")
            cell_coord = f"{openpyxl.utils.get_column_letter(picture_col_idx + 1)}{row_num}"
            if cell_coord in images:
                df.at[idx, "PICTURE_BASE64"] = images[cell_coord]
        except:
            pass
    
    return df, images, ws

def create_status_badge(status):
    """Tạo badge HTML cho status"""
    if status == "Completed":
        return '<span class="status-completed">✓ Completed</span>'
    elif status == "Working":
        return '<span class="status-working">⚙️ Working</span>'
    elif status == "Delay":
        return '<span class="status-delay">⚠️ Delay</span>'
    elif status == "New Task":
        return '<span class="status-newtask">🆕 New Task</span>'
    return status

# === Tiêu đề chính ===
st.title("📋 Task Dashboard")
st.markdown("---")

# === Sidebar - Upload file ===
with st.sidebar:
    st.header("⚙️ Cấu hình")
    uploaded_file = st.file_uploader(
        "Upload Excel File",
        type=["xlsx", "xls"],
        help="Upload file Excel với header ở dòng 3"
    )
    
    if uploaded_file is not None:
        st.success("✅ File đã được tải lên!")
        
        # Load dữ liệu
        try:
            df, images, ws = load_and_process_data(uploaded_file)
            
            st.markdown("---")
            st.subheader("📊 Thống kê tổng quan")
            
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Tổng số task", len(df))
                st.metric("Task có hình ảnh", len(df[df["PICTURE_BASE64"] != ""]))
            
            with col2:
                st.metric("Completed", len(df[df["STATUS"] == "Completed"]))
                st.metric("Delay", len(df[df["STATUS"] == "Delay"]))
            
            st.markdown("---")
            st.subheader("🔍 Lọc dữ liệu")
            
            # Filter theo STATUS
            status_filter = st.multiselect(
                "Chọn STATUS",
                options=["All"] + sorted(df["STATUS"].unique().tolist()),
                default=["All"]
            )
            
            # Filter theo ngày
            date_range = st.date_input(
                "Lọc theo START DATE",
                value=None,
                help="Để trống để hiển thị tất cả"
            )
            
        except Exception as e:
            st.error(f"❌ Lỗi khi đọc file: {str(e)}")
            st.stop()
    else:
        st.info("👆 Vui lòng upload file Excel để bắt đầu")
        st.stop()

# === Main content ===
if uploaded_file is not None:
    
    # Áp dụng filter
    df_filtered = df.copy()
    
    if "All" not in status_filter and len(status_filter) > 0:
        df_filtered = df_filtered[df_filtered["STATUS"].isin(status_filter)]
    
    if date_range:
        if isinstance(date_range, tuple) and len(date_range) == 2:
            start_date, end_date = date_range
            df_filtered = df_filtered[
                (df_filtered["START DATE"] >= pd.Timestamp(start_date)) &
                (df_filtered["START DATE"] <= pd.Timestamp(end_date))
            ]
    
    # === Tab layout ===
    tab1, tab2, tab3 = st.tabs(["📊 Biểu đồ", "📋 Bảng dữ liệu", "🖼️ Hình ảnh"])
    
    # === TAB 1: Biểu đồ ===
    with tab1:
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("Tỷ lệ STATUS các Task")
            status_counts = df_filtered["STATUS"].value_counts()
            if not status_counts.empty:
                fig_pie = px.pie(
                    df_filtered,
                    names="STATUS",
                    color="STATUS",
                    color_discrete_map={
                        "Completed": "green",
                        "Working": "orange",
                        "Delay": "red",
                        "New Task": "blue"
                    }
                )
                fig_pie.update_traces(textinfo='percent+label', pull=[0.05]*len(status_counts))
                st.plotly_chart(fig_pie, use_container_width=True)
            else:
                st.info("Không có dữ liệu để hiển thị")
        
        with col2:
            st.subheader("Phân bố theo tháng")
            df_with_dates = df_filtered[df_filtered["START DATE"].notna()].copy()
            if not df_with_dates.empty:
                df_with_dates["month"] = df_with_dates["START DATE"].dt.strftime("%Y-%m")
                df_summary = df_with_dates.groupby(["month", "STATUS"]).size().reset_index(name="count")
                
                # Tạo DataFrame đầy đủ
                all_months = sorted(df_summary["month"].unique())
                all_statuses = ["Completed", "Working", "New Task", "Delay"]
                
                full_data = []
                for month in all_months:
                    for status in all_statuses:
                        existing = df_summary[(df_summary["month"] == month) & (df_summary["STATUS"] == status)]
                        if not existing.empty:
                            full_data.append({"month": month, "STATUS": status, "count": existing["count"].values[0]})
                        else:
                            full_data.append({"month": month, "STATUS": status, "count": 0})
                
                df_full = pd.DataFrame(full_data)
                
                # Tạo bar chart
                fig_bar = go.Figure()
                colors = {
                    "Completed": "green",
                    "Working": "orange",
                    "New Task": "blue",
                    "Delay": "red"
                }
                
                for status in all_statuses:
                    df_status = df_full[df_full["STATUS"] == status]
                    fig_bar.add_trace(go.Bar(
                        x=df_status["month"],
                        y=df_status["count"],
                        name=status,
                        marker_color=colors.get(status, "gray"),
                        text=df_status["count"],
                        textposition='outside',
                        textfont=dict(size=10),
                    ))
                
                fig_bar.update_layout(
                    barmode='group',
                    xaxis=dict(tickformat="%Y-%m", type='category'),
                    hovermode='x unified',
                    height=400,
                    showlegend=True,
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
                )
                st.plotly_chart(fig_bar, use_container_width=True)
            else:
                st.info("Không có dữ liệu ngày tháng để hiển thị")
    
    # === TAB 2: Bảng dữ liệu ===
    with tab2:
        st.subheader(f"Danh sách Task ({len(df_filtered)} tasks)")
        
        # Tạo DataFrame hiển thị
        df_display = df_filtered.copy()
        df_display["START DATE"] = df_display["START DATE"].dt.strftime("%m/%d/%Y")
        df_display["DUE DATE"] = df_display["DUE DATE"].dt.strftime("%m/%d/%Y")
        df_display = df_display.fillna("")
        
        # Chọn các cột hiển thị
        display_cols = ["TASK", "Requester", "START DATE", "DUE DATE", "CONFIRM FROM BARON", "STATUS"]
        df_show = df_display[display_cols].copy()
        
        # Tạo HTML table với status có màu
        html_table = "<table style='width:100%; border-collapse: collapse;'>"
        html_table += "<thead><tr style='background-color: #4CAF50; color: white;'>"
        for col in display_cols:
            html_table += f"<th style='padding: 10px; border: 1px solid #ddd;'>{col}</th>"
        html_table += "</tr></thead><tbody>"
        
        for idx, row in df_show.iterrows():
            html_table += "<tr>"
            for col in display_cols:
                value = str(row[col]) if row[col] != "" else ""
                if col == "STATUS":
                    value = create_status_badge(value)
                    html_table += f"<td style='padding: 8px; border: 1px solid #ddd; text-align: center;'>{value}</td>"
                else:
                    html_table += f"<td style='padding: 8px; border: 1px solid #ddd;'>{value}</td>"
            html_table += "</tr>"
        
        html_table += "</tbody></table>"
        st.markdown(html_table, unsafe_allow_html=True)
        
        # Download button
        st.markdown("---")
        csv = df_show.to_csv(index=False).encode('utf-8-sig')
        st.download_button(
            label="📥 Download CSV",
            data=csv,
            file_name=f"task_dashboard_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv",
        )
    
    # === TAB 3: Hình ảnh ===
    with tab3:
        st.subheader("Thư viện hình ảnh Task")
        
        # Filter tasks có hình ảnh
        df_with_images = df_filtered[df_filtered["PICTURE_BASE64"] != ""].copy()
        
        if len(df_with_images) > 0:
            st.info(f"Tìm thấy {len(df_with_images)} task có hình ảnh")
            
            # Hiển thị grid hình ảnh
            cols_per_row = 3
            rows = (len(df_with_images) + cols_per_row - 1) // cols_per_row
            
            for row_idx in range(rows):
                cols = st.columns(cols_per_row)
                for col_idx in range(cols_per_row):
                    img_idx = row_idx * cols_per_row + col_idx
                    if img_idx < len(df_with_images):
                        task_row = df_with_images.iloc[img_idx]
                        with cols[col_idx]:
                            st.markdown(f"**{task_row['TASK']}**")
                            st.markdown(f"*Status: {create_status_badge(task_row['STATUS'])}*", unsafe_allow_html=True)
                            
                            # Hiển thị hình ảnh
                            img_data = task_row["PICTURE_BASE64"]
                            if img_data:
                                img_html = f'<img src="data:image/png;base64,{img_data}" style="width:100%; border-radius:8px; border:1px solid #ddd;"/>'
                                st.markdown(img_html, unsafe_allow_html=True)
                            
                            st.markdown("---")
        else:
            st.warning("Không có task nào có hình ảnh trong bộ lọc hiện tại")
    
    # === Footer ===
    st.markdown("---")
    st.markdown(
        f"<div style='text-align: center; color: gray; padding: 10px;'>"
        f"Dashboard cập nhật lần cuối: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        f"</div>",
        unsafe_allow_html=True
    )
